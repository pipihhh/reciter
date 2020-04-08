from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from copy import copy
from utils import set_row_and_col
from conf import CONF

RESORT_SHEETS = "ydspc"


def resort_all_words():
    wb = load_workbook(CONF["pathOfExcel"])
    cell_list = []
    diff_set = set()
    target_sheet = None
    for sheet in wb:
        if sheet.title == RESORT_SHEETS:
            target_sheet = sheet
            for col in range(1, sheet.max_column + 1, 2):
                for row in range(1, sheet.max_row + 1):
                    cell = sheet.cell(row, col)
                    if cell.value is not None and cell.value not in diff_set:
                        diff_set.add(cell.value)
                        cell_list.append({
                            "value": cell.value,
                            "comment": copy(cell.comment),
                            # "fill": PatternFill(
                            #     patternType=cell.fill.patternType,
                            #     bgColor=cell.fill.bgColor.rgb
                            # )
                            "fill": copy(cell.fill)
                        })
    if target_sheet is not None:
        row, col = 0, 1
        for cell in cell_list:
            row, col = set_row_and_col(row, col)
            target_sheet.cell(row, col, cell["value"])
            target_sheet.cell(row, col).comment = cell["comment"]
            target_sheet.cell(row, col).fill = cell["fill"]
        while True:
            row, col = set_row_and_col(row, col)
            if row <= target_sheet.max_row and col <= target_sheet.max_column:
                cell = target_sheet.cell(row, col)
                cell.value = None
                cell.comment = None
                cell.fill = PatternFill()
            else:
                break
    wb.save(CONF["pathOfExcel"])
    wb.close()


if __name__ == '__main__':
    resort_all_words()
