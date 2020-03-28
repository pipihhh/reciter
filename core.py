from openpyxl import load_workbook
from openpyxl.comments.comments import Comment
import random
from command_handler import CommandHandler

CONF = {
    "pathOfExcel": "C:\\Users\\qq312\\Desktop\\生僻词总结本子.xlsx",
    "nameOfSheets": ["llyc", "ydspc", "ydcz"],
    "times": 1
}


class OneSheetDB(object):
    def __init__(self, word_dict, word_index):
        self._word_dict = word_dict
        self._word_index = word_index
        self._difficulties = {}
        self._former_word = None

    def get_translation(self, word):
        return self._word_dict[word][0].content

    def get_random(self):
        rand_index = random.randint(0, len(self._word_index)-1)
        # self._former_word = (self._word_index[rand_index], self._word_dict[self._word_index[rand_index]])
        return self._word_index[rand_index]

    def add(self, word):
        self._word_dict[word][1] += 1

    def divide(self, word):
        self._word_dict[word][1] -= 1

    def get_number(self, word):
        return self._word_dict[word][1]

    def pop(self, word):
        try:
            self._former_word = (word, self._word_dict[word])
            self._word_dict.pop(word)
            self._word_index.remove(word)
        except KeyError:
            print("the word is not in the wordsDb")

    def is_empty(self):
        return False if len(self._word_index) > 0 else True

    def add_log(self, word):
        self._difficulties[word] = self.get_translation(word)

    def add_former(self):
        if self._former_word is not None:
            self._difficulties[self._former_word[0]] = self._former_word[1][0].content
            self._word_index.append(self._former_word[0])
            self._word_dict[self._former_word[0]] = self._former_word[1]

    def __len__(self):
        return len(self._word_index)

    def finish(self, sheet):
        """
        将不会的单词存入sheet 不会存入重复的单词
        :param sheet:
        :return:
        """
        # print(self._difficulties)
        col = 1
        row = 1
        for word, comment in self._difficulties.items():
            flag = True
            while True:
                if row > 25:
                    row = 1
                    col += 2
                data = sheet.cell(row, col)
                if data.value is not None and data.comment is not None:
                    row += 1
                    if data.value in self._difficulties:
                        self._difficulties[data.value] = None
                        if data.value == word:
                            flag = False
                    continue
                else:
                    break
            if flag and comment is not None:
                data.value = word
                data.comment = Comment(text=comment, author="Lee Mist")
                row += 1


# def init_word_dict(sheet):
#     word_dict = {}
#     word_index = []
#     for col in range(1, sheet.max_column+1, 2):
#         for row in range(1, sheet.max_row+1):
#             data = sheet.cell(row, col)
#             if data.value in word_dict or data.value is None:
#                 continue
#             word_index.append(data.value)
#             word_dict[data.value] = [data.comment, 0]
#     return OneSheetDB(word_dict, word_index)


def init_all_llyc():
    wb = load_workbook(CONF["pathOfExcel"])
    word_dict = {}
    word_index = []
    sheets_set = {"llyc2", }
    for sheet in wb:
        if sheet.title in sheets_set:
            for col in range(1, sheet.max_column + 1, 2):
                for row in range(1, sheet.max_row + 1):
                    data = sheet.cell(row, col)
                    if data.value in word_dict or data.value is None:
                        continue
                    word_index.append(data.value)
                    word_dict[data.value] = [data.comment, 0]
    return OneSheetDB(word_dict, word_index)


if __name__ == '__main__':
    """
        核心的背单词的文件 只需要运行此文件即可
        y为认识此单词 n为不认识此单词 eof为退出
        error为rollback上一个单词入单词池
        excel需要提取的sheet在108行的sheets_set来指定 可以指定多个set
        在背诵过程中error过或n过的单词在程序正常结束后会被记入此文件的difficulties的sheet中
        需要人为创建这个名为difficulties的sheet或者自行指定一个名称
    """
    # wb = load_workbook(CONF["pathOfExcel"])
    # sheet_word = wb[CONF["nameOfSheets"][0]]
    wordsDb = init_all_llyc()
    count = CONF["times"] - 1
    handler = CommandHandler()
    # wb.close()
    while not wordsDb.is_empty():
        word = wordsDb.get_random()
        print(f"{word}({len(wordsDb)})")
        ans = input()
        if ans.upper() == "ERROR":
            wordsDb.add_former()
            ans = input()
        number = wordsDb.get_number(word)
        translation = wordsDb.get_translation(word)
        if ans.upper() == "EOF":
            break
        elif ans.lower() == "y" and number < count:
            wordsDb.add(word)
        elif ans.lower() == "y" and number >= count:
            wordsDb.pop(word)
        elif ans.lower() == "n":
            wordsDb.add_log(word)
        else:
            handler.handle(ans, wordsDb)
        print("\n答案:\n")
        print(translation)
        print("\n")
    wb = load_workbook(CONF["pathOfExcel"])
    difficulties = wb["difficulties"]
    wordsDb.finish(difficulties)
    wb.save(CONF["pathOfExcel"])
    wb.close()
