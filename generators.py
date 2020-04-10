from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments.comments import Comment
from utils import set_row_and_col, parse_letter_to_num, get_settings, get_index_by_color
import random
from conf import *


class RandomWordGenerator(object):
    def __init__(self, wb):
        self._word_dict = {}
        self._word_index = []
        self._difficulties = {}
        self._wb = wb

    def get_translation(self, word):
        return self._word_dict[word]["cell"].comment.content

    def get_random(self):
        rand_index = random.randint(0, len(self._word_index) - 1)
        return self._word_index[rand_index]

    def pop(self, word):
        try:
            self._word_index.remove(word)
        except Exception:
            import traceback
            print(traceback.format_exc())

    def set_work_seconds(self, word, seconds):
        if word in self._word_dict:
            self._word_dict[word]["work_out_seconds"] = seconds

    def is_empty(self) -> bool:
        return False if len(self._word_index) > 0 else True

    def add_log(self, word):
        if word in self._difficulties:
            self._difficulties[word]["forget_times"] += 1
            return
        self._difficulties[word] = {"translation": self.get_translation(word), "forget_times": 1}

    def __len__(self):
        return len(self._word_index)

    def finish(self):
        """
        将不会的单词存入sheet 不会存入重复的单词
        :return:
        """
        sheet = self._wb["difficulties"]
        col = 1
        row = 0
        word_dict = {}
        while True:
            row, col = set_row_and_col(row, col)
            cell = sheet.cell(row, col)
            if cell.value is not None:
                word_dict[cell.value] = cell
            else:
                break
        row -= 1
        for word, d in self._word_dict.items():
            cell = d["cell"]
            if word in self._difficulties:
                if word in word_dict:
                    self._set_bg_color(cell, self._difficulties[word]["forget_times"] + 1)
                else:
                    row, col = set_row_and_col(row, col)
                    self._set_bg_color(cell, self._difficulties[word]["forget_times"])
                    sheet.cell(row, col).value = word
                    sheet.cell(row, col).comment = Comment(
                        text=self._difficulties[word]["translation"], author="Lee Mist"
                    )
            else:
                if len(self) <= 0:
                    if get_settings().difficulty_mode is True:
                        self.reduce_color_level(cell.value)
                        continue
                    self._set_bg_color(cell, 0)
        self._wb.save(CONF["pathOfExcel"])
        self._wb.close()

    def _set_bg_color(self, cell, forget_times):
        if forget_times < CONF["fill_threshold"]:
            cell.fill = PatternFill()
            return
        forget_times = forget_times - CONF["fill_threshold"] + 2
        color = COLOR_DICT.get(forget_times, RED)
        cell.fill = PatternFill(fill_type="solid", fgColor=color)

    def append(self, cell):
        if cell.value not in self._word_dict and cell.value is not None:
            self._word_dict[cell.value] = {"cell": cell, "work_out_seconds": None}
            self._word_index.append(cell.value)

    def up_color_level(self, w):
        cell = self._word_dict[w]["cell"]
        color_idx = get_index_by_color(cell.fill.fgColor.rgb)
        self._set_bg_color(cell, color_idx + CONF["fill_threshold"] - 1)

    def reduce_color_level(self, w):
        cell = self._word_dict[w]["cell"]
        color_idx = get_index_by_color(cell.fill.fgColor.rgb)
        self._set_bg_color(cell, color_idx + CONF["fill_threshold"] - 3)

    @classmethod
    def init_all_words(cls):
        wb = load_workbook(CONF["pathOfExcel"])
        word_db = cls(wb)
        settings = get_settings()
        for sheet in wb:
            if sheet.title in settings.sheets and settings.sheets[sheet.title]:
                for col in range(1, sheet.max_column + 1, 2):
                    for row in range(1, sheet.max_row + 1):
                        word_db.append(sheet.cell(row, col))
        return word_db


class PartlyWordGenerator(RandomWordGenerator):
    @classmethod
    def init_all_words(cls):
        wb = load_workbook(CONF["pathOfExcel"])
        obj = cls(wb)
        settings = get_settings()
        for sheet in wb:
            if sheet.title in settings.partly_dict and settings.partly_dict[sheet.title]["flag"]:
                if isinstance(settings.partly_dict[sheet.title], list):
                    for d in settings.partly_dict[sheet.title]:
                        row, col = 0, parse_letter_to_num(d["start"])
                        end = parse_letter_to_num(d["end"])
                        PartlyWordGenerator.fill(sheet, row, col, end, obj)
                else:
                    row, col = 0, parse_letter_to_num(settings.partly_dict[sheet.title]["start"])
                    end = parse_letter_to_num(settings.partly_dict[sheet.title]["end"])
                    PartlyWordGenerator.fill(sheet, row, col, end, obj)
                # while True:
                #     row, col = set_row_and_col(row, col)
                #     if col > end:
                #         break
                #     cell = sheet.cell(row, col)
                #     obj.append(cell)
        return obj

    @staticmethod
    def fill(sheet, row, col, end, obj):
        """
        填充单元格
        :param sheet: 要被读取的sheet
        :param row: 起始位置的row
        :param col: 起始位置的col
        :param end: 结束位置的 用来判断
        :param obj: 填充进的对象
        :return:
        """
        while True:
            row, col = set_row_and_col(row, col)
            if col > end:
                break
            cell = sheet.cell(row, col)
            obj.append(cell)


class EmphasizedWordGenerator(RandomWordGenerator):
    @classmethod
    def init_all_words(cls):
        wb = load_workbook(CONF["pathOfExcel"])
        obj = cls(wb)
        for sheet in wb:
            row, col = 0, 1
            while True:
                row, col = set_row_and_col(row, col)
                cell = sheet.cell(row, col)
                if cell.fill.fgColor.rgb != "00000000":
                    obj.append(cell)
                if col > 25:
                    break
        return obj


class RandomlyChooseGenerator(RandomWordGenerator):
    @classmethod
    def init_all_words(cls):
        wb = load_workbook(CONF["pathOfExcel"])
        word_dict = {}
        word_list = []
        for sheet in wb:
            for col in range(1, sheet.max_column + 1, 2):
                for row in range(1, sheet.max_row + 1):
                    cell = sheet.cell(row, col)
                    if cell.value is not None and cell.value not in word_dict:
                        word_dict[cell.value] = cell
                        word_list.append(cell.value)
        offset = get_settings().random_count
        obj = cls(wb)
        if offset > len(word_list):
            RandomlyChooseGenerator.fill(obj, word_dict, word_list)
        else:
            from random import shuffle
            shuffle(word_list)
            RandomlyChooseGenerator.fill(obj, word_dict, word_list[:offset])
        return obj

    @staticmethod
    def fill(obj, word_dict, word_list):
        for word in word_list:
            obj.append(word_dict[word])
