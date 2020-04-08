from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from utils import get_settings
from conf import CONF


def diff_sheets(work_book):
    words_pool = {}
    error_list = []
    settings = get_settings()
    for sheet in work_book:
        if sheet.title not in settings.sheets:
            continue
        for col in range(1, sheet.max_column+1, 2):
            for row in range(1, sheet.max_row+1):
                data = sheet.cell(row, col)
                if data.value is None and data.comment is None:
                    continue
                if data.value not in words_pool:
                    words_pool[data.value] = f"{sheet.title}:{get_column_letter(col)}:{row}"
                else:
                    print(f"{data.value}:({sheet.title}:{get_column_letter(col)}:{row}->{words_pool[data.value]})")
                    data.value = None
                    data.comment = None


if __name__ == '__main__':
    ws = load_workbook(CONF["pathOfExcel"])
    diff_sheets(ws)
    ws.save(CONF["pathOfExcel"])
