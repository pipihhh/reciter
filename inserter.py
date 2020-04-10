from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font
from generators import RandomWordGenerator
from utils import get_settings, set_row_and_col, get_index_by_color, merge
from conf import CONF


class Insert(RandomWordGenerator):
    settings = get_settings()

    @classmethod
    def init_all_words(cls):
        wb = load_workbook(CONF["pathOfExcel"])
        word_db = cls(wb)
        for sheet in wb:
            for col in range(1, sheet.max_column + 1, 2):
                for row in range(1, sheet.max_row + 1):
                    word_db.append(sheet.cell(row, col))
        col, row = 1, 0
        sheet = wb[cls.settings.insert["sheet_name"]]
        while True:
            row, col = set_row_and_col(row, col)
            if sheet.cell(row, col).value is not None:
                continue
            break
        word_db.col = col
        word_db.row = row
        return word_db

    def __init__(self, wb):
        RandomWordGenerator.__init__(self, wb)
        self.col = None
        self.row = None
        self.sheet_name = self.settings.insert["sheet_name"]

    def __contains__(self, item):
        return item in self._word_dict

    def save(self):
        self._wb.save(CONF["pathOfExcel"])
        self._wb.close()

    def insert(self, w, trans):
        sheet = self._wb[self.sheet_name]
        while True:
            cell = sheet.cell(self.row, self.col)
            self.row, self.col = set_row_and_col(self.row, self.col)
            if cell.value is not None:
                continue
            break
        cell.value = w
        cell.comment = Comment(text=trans, author="Lee Mist")
        cell.font = Font(self.settings.insert["font"], size=18, color="000000")

    def append_mode(self, w):
        trans = self.initial_translation()
        trans = merge(self._word_dict[w]["cell"].comment.content, trans)
        if verify(w, trans):
            self._word_dict[w]["cell"].comment = Comment(text=trans, author="Lee Mist")
            self.up_color_level(w)

    @staticmethod
    def initial_translation():
        translation = []
        while True:
            t = input("translation:")
            if t.upper() == "ERROR":
                translation.pop()
                continue
            if t.upper() == "EOF":
                break
            else:
                translation.append(t.strip())
        return "\n".join(translation)


def verify(w, trans):
    if Insert.settings.insert["strict_mode"]:
        r = input(f"{w}的汉译将被修改为:\n{trans}\n[Y or N]:")
        if r.upper() == "N":
            return False
        return True
    print(f"{w}的汉译将被修改为:\n{trans}")
    return True


if __name__ == '__main__':
    db = Insert.init_all_words()
    while True:
        argv = input("单词:").split()
        if len(argv) > 1:
            word, options = argv[0], argv[1:]
        else:
            word, options = argv[0], []
        if word.upper() == "EOF":
            break
        if "-a" in options:
            db.append_mode(word)
            continue
        if word in db:
            print("此单词已存在")
            db.up_color_level(word)
            continue
        trans = Insert.initial_translation()
        if verify(word, trans):
            db.insert(word, trans)
    db.save()
